from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment

excel_sheet = '#TEMP'
excel_eoc = '#END_OF_COL'
excel_eor = '#END_OF_ROW'
excel_label_prefix = '#'


def read_template(params):
    excel_params = params['excel']
    template_path = excel_params['template_path']
    align = Alignment(horizontal='center', vertical='center')
    workbook = load_workbook(filename=template_path)
    sheet = workbook[excel_sheet]
    template_dict = {}
    row = 1
    loop_col = True
    while loop_col:
        col = 1
        loop_row = True
        while loop_row:
            value = str(sheet.cell(row=row, column=col).value).strip()
            if value == excel_eoc:
                loop_col = False
                sheet.cell(row=row, column=col, value='')
                sheet.cell(row=row, column=col).alignment = align
            elif value == excel_eor:
                loop_row = False
                sheet.cell(row=row, column=col, value='')
                sheet.cell(row=row, column=col).alignment = align
            else:
                if value.startswith(excel_label_prefix):
                    key = value.replace(excel_label_prefix, '')
                    template_dict.__setitem__(key, {'row': row, 'col': col})
            col += 1
        row += 1
    excel_params['template_dict'] = template_dict
    excel_params['workbook'] = workbook
    return params


def read_testcases(params):
    def rt_per_key(k, s, td):
        kwd = k.replace('↑', '').replace('↓', '').replace('←', '').replace('→', '')
        if kwd == k:
            cp = s.cell(row=td[k]['row'], column=td[k]['col']).value
        else:
            direction = k.replace(kwd, '')
            cp = []
            if direction == '↑':
                base_row = td[k]['row']
                while True:
                    cell_value = s.cell(row=base_row, column=td[k]['col']).value
                    if cell_value not in [excel_eoc, excel_eor]:
                        cp.append(cell_value)
                        base_row -= 1
                    else:
                        break
            elif direction == '↓':
                base_row = td[k]['row']
                while True:
                    cell_value = s.cell(row=base_row, column=td[k]['col']).value
                    if cell_value not in [excel_eoc, excel_eor]:
                        cp.append(cell_value)
                        base_row += 1
                    else:
                        break
            elif direction == '←':
                base_col = td[k]['col']
                while True:
                    cell_value = s.cell(row=td[k]['row'], column=base_col).value
                    if cell_value not in [excel_eoc, excel_eor]:
                        cp.append(cell_value)
                        base_col -= 1
                    else:
                        break
            elif direction == '→':
                base_col = td[k]['col']
                while True:
                    cell_value = s.cell(row=td[k]['row'], column=base_col).value
                    if cell_value not in [excel_eoc, excel_eor]:
                        cp.append(cell_value)
                        base_col += 1
                    else:
                        break
        return cp

    excel_params = params['excel']
    template_dict = excel_params['template_dict']
    testcase_path = excel_params['testcase_path']
    workbook = load_workbook(filename=testcase_path)
    sheet = workbook[excel_sheet]
    td_copy = template_dict.copy()
    case_params = {}
    for key in td_copy.keys():
        case_param = rt_per_key(key, sheet, template_dict)
        key_without_direction = key.replace('↑', '').replace('↓', '').replace('←', '').replace('→', '')
        case_params.__setitem__(key_without_direction, case_param)
    if 'case_params' in excel_params.keys():
        for key in case_params.keys():
            excel_params['case_params'][key] = case_params[key]
    else:
        excel_params['case_params'] = case_params
    return params


def write_summary(params):
    def ws_per_key(dsd, k, s, td, d):
        align = Alignment(horizontal='center', vertical='center')
        key_without_direction = k.replace('↑', '').replace('↓', '').replace('←', '').replace('→', '')
        data_src = dsd[key_without_direction]
        if type(data_src) is not list:
            s.cell(row=td[k]['row'], column=td[k]['col'], value=str(data_src))
            ws_recursive(row=td[k]['row'], col=td[k]['col'], dsd=dsd, element=data_src, s=s, td=td, d=d)
            s.cell(row=td[k]['row'], column=td[k]['col']).alignment = align
        else:
            direction = k.replace(key_without_direction, '')
            if direction == '↑':
                base_row = td[k]['row']
                for element in data_src:
                    s.cell(row=base_row, column=td[k]['col'], value=str(element))
                    ws_recursive(row=base_row, col=td[k]['col'], dsd=dsd, element=element, s=s, td=td, d=d)
                    s.cell(row=base_row, column=td[k]['col']).alignment = align
                    base_row -= 1
                    if base_row <= 0:
                        break
            elif direction == '↓':
                base_row = td[k]['row']
                for element in data_src:
                    s.cell(row=base_row, column=td[k]['col'], value=str(element))
                    ws_recursive(row=base_row, col=td[k]['col'], dsd=dsd, element=element, s=s, td=td, d=d)
                    s.cell(row=base_row, column=td[k]['col']).alignment = align
                    base_row += 1
            elif direction == '←':
                base_col = td[k]['col']
                for element in data_src:
                    s.cell(row=td[k]['row'], column=base_col, value=str(element))
                    ws_recursive(row=td[k]['row'], col=base_col, dsd=dsd, element=element, s=s, td=td, d=d)
                    s.cell(row=td[k]['row'], column=base_col).alignment = align
                    base_col -= 1
                    if base_col <= 0:
                        break
            elif direction == '→':
                base_col = td[k]['col']
                for element in data_src:
                    s.cell(row=td[k]['row'], column=base_col, value=str(element))
                    ws_recursive(row=td[k]['row'], col=base_col, dsd=dsd, element=element, s=s, td=td, d=d)
                    s.cell(row=td[k]['row'], column=base_col).alignment = align
                    base_col += 1

    def ws_recursive(row, col, dsd, element, s, td, d):
        if d > 0 and str(element).startswith(excel_label_prefix):
            d -= 1
            element = str(element).replace(excel_label_prefix, '')
            dsd.__setitem__(str(element), {'row': row, 'col': col})
            ws_per_key(
                dsd=dsd, k=str(element), s=s,
                td=td, d=d
            )

    excel_params = params['excel']
    template_dict = excel_params['template_dict']
    workbook = excel_params['workbook']
    data_src_dict = excel_params['data_src_dict']

    sheet = workbook[excel_sheet]
    td_copy = template_dict.copy()
    if 'depth' in excel_params.keys():
        depth = excel_params['depth']
    else:
        depth = 0
    for key in td_copy.keys():
        ws_per_key(dsd=data_src_dict, k=key, s=sheet, td=template_dict, d=depth)
    excel_params['workbook'] = workbook
    if 'save_path' in excel_params.keys():
        save_path = excel_params['save_path']
        finished = False
        while not finished:
            try:
                workbook.save(filename=save_path)
                finished = True
            except Exception as e:
                _ = e
    return params
